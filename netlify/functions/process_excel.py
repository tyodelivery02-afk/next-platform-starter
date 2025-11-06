# netlify/functions/process_excel.py
import json
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

def handler(event, context):
    try:
        # 检查 HTTP 方法
        if event.get('httpMethod') != 'POST':
            return {
                'statusCode': 405,
                'body': json.dumps({'error': 'Method not allowed'})
            }
        
        # 解析请求体
        body = json.loads(event.get('body', '{}'))
        
        # 获取 base64 编码的文件和统计数据
        file_base64 = body.get('file')
        stats_data = body.get('statsData')
        
        if not file_base64 or not stats_data:
            return {
                'statusCode': 400,
                'body': json.dumps({'error': '缺少必要参数: file 或 statsData'})
            }
        
        # 解码文件
        file_data = base64.b64decode(file_base64)
        
        # 打开 Excel（从内存中）
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active

        # 样式定义
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        bold_font = Font(name="Arial", size=16, bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        fixed_col_width = 10

        def style_cell(cell, val):
            cell.fill = yellow_fill
            cell.alignment = center_align
            cell.font = bold_font
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.number_format = '0'
            else:
                cell.number_format = '@'
            cell.value = val
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = fixed_col_width

        # ---------------- 3. 新增列 & 写标题 ----------------
        start_col = 15  # O列
        stats_cols = stats_data[0]  # 标题行
        insert_count = len(stats_cols)

        # 插入列
        ws.insert_cols(start_col, amount=insert_count)

        # 写标题（第4行）
        style_cell(ws.cell(row=4, column=start_col), "全体")
        for i, header in enumerate(stats_cols[1:]):
            style_cell(ws.cell(row=4, column=start_col + 1 + i), header)

        # ---------------- 4. 按マスタ番号匹配写入数据 & 隐藏行 ----------------
        stats_rows = stats_data[1:-2]
        total_rows = stats_data[-2:]

        for row_idx in range(5, ws.max_row + 1):
            master_no = ws.cell(row_idx, column=7).value  # G列
            match = next((r for r in stats_rows if r[0] == master_no), None)

            if match:
                # 计算全体 = 第2列 + 第3列
                try:
                    val2 = float(match[1]) if match[1] not in [None, ""] else 0
                except:
                    val2 = 0
                try:
                    val3 = float(match[2]) if match[2] not in [None, ""] else 0
                except:
                    val3 = 0
                total_val = val2 + val3
                style_cell(ws.cell(row_idx, column=start_col), total_val)

                # 写入原 statsData 数据
                for i, val in enumerate(match[1:]):
                    style_cell(ws.cell(row_idx, column=start_col + 1 + i), val)

                ws.row_dimensions[row_idx].hidden = False
            else:
                # 如果整行完全为空，不隐藏；否则隐藏
                row_cells = ws[row_idx]
                if all(cell.value in [None, ""] for cell in row_cells):
                    ws.row_dimensions[row_idx].hidden = False
                else:
                    ws.row_dimensions[row_idx].hidden = True

        # ---------------- 5. 写入总计行 ----------------
        def is_row_empty(row):
            for cell in row:
                if cell.value not in [None, ""]:
                    return False
            return True

        # 找到最后有数据且未隐藏行
        last_data_row = 4
        for row_idx in range(5, ws.max_row + 1):
            row = ws[row_idx]
            if not is_row_empty(row):
                last_data_row = row_idx

        for idx, t_row in enumerate(total_rows):
            new_row_idx = last_data_row + 1 + idx
            ws.cell(row=new_row_idx, column=14, value=t_row[0])
            
            # 计算全体列
            if idx == 0:  # 総計1
                total_val = float(total_rows[1][1]) + float(total_rows[1][2]) if total_rows[1][1] not in [None,""] and total_rows[1][2] not in [None,""] else 0
            else:
                try:
                    val2 = float(t_row[1]) if t_row[1] not in [None, ""] else 0
                except:
                    val2 = 0
                try:
                    val3 = float(t_row[2]) if t_row[2] not in [None, ""] else 0
                except:
                    val3 = 0
                total_val = val2 + val3

            style_cell(ws.cell(new_row_idx, column=start_col), total_val)

            # 写入原总计数据
            for j, val in enumerate(t_row[1:]):
                style_cell(ws.cell(new_row_idx, column=start_col + 1 + j), val)

        # ---------------- 6. 代理店计算（竖向写入） ----------------
        agent_map = {}
        for row_idx in range(5, ws.max_row + 1):
            if ws.row_dimensions[row_idx].hidden:
                continue
            agent = ws.cell(row=row_idx, column=8).value  # H列
            tokyo_val = ws.cell(row=row_idx, column=start_col + 1).value
            try:
                tokyo_val = float(tokyo_val) if tokyo_val not in [None, ""] else 0
            except:
                tokyo_val = 0
            agent_map[agent] = agent_map.get(agent, 0) + tokyo_val

        agent_div = {"CAINIAO-E": 450, "TEMU": 250, "CNE": 180}
        insert_row_num = last_data_row + 1 + len(total_rows)

        # 写标题 O列
        cell_title = ws.cell(row=insert_row_num, column=15, value="代理店")
        cell_title.fill = orange_fill
        cell_title.alignment = center_align
        cell_title.font = bold_font
        cell_title.border = thin_border

        # 竖向写入
        start_row = insert_row_num + 1
        for idx, (agent, total) in enumerate(agent_map.items()):
            div = agent_div.get(agent)
            if div:
                val = (total // div) + (1 if total % div else 0)
            else:
                val = "計算外"

            # 代理店名竖向写 O列
            cell_agent = ws.cell(row=start_row + idx, column=15, value=agent)
            cell_agent.fill = orange_fill
            cell_agent.alignment = center_align
            cell_agent.font = bold_font
            cell_agent.border = thin_border

            # 计算结果竖向写 Q列
            cell_val = ws.cell(row=start_row + idx, column=17, value=val)
            cell_val.fill = orange_fill
            cell_val.alignment = center_align
            cell_val.font = bold_font
            cell_val.border = thin_border

        # ---------------- 7. 保存 Excel 到内存 ----------------
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 转换为 base64 返回
        output_base64 = base64.b64encode(output.read()).decode('utf-8')

        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename=filled.xlsx'
            },
            'body': output_base64,
            'isBase64Encoded': True
        }
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error: {str(e)}")
        print(f"Traceback: {error_detail}")
        
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': str(e),
                'type': type(e).__name__,
                'traceback': error_detail
            })
        }